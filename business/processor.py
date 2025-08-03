import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Callable
import tempfile
import shutil

import gspread
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io

from .config import Config, load_config, BASE_DIR
from .logic.sheet_utils import copy_sheet_data, clear_column_cache


class ExcelToGoogleSheets:
    """Класс для копирования данных из Excel в Google Таблицы."""

    def __init__(self, config_path: str = "config.yaml"):
        path = Path(config_path)
        if not path.is_absolute():
            path = BASE_DIR / path
        self.config_path = str(path)
        self.config = load_config(self.config_path)
        self.logger = self._setup_logger()
        self.gc = None
        self.google_sheet = None
        self._google_creds = None
        self._drive_service = None

    def _setup_logger(self) -> logging.Logger:
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        return logger

    def extract_sheet_id_from_url(self, url: str) -> str:
        """Извлечение ID таблицы из URL Google Sheets."""
        import re
        patterns = [
            r'/spreadsheets/d/([a-zA-Z0-9-_]+)',
            r'id=([a-zA-Z0-9-_]+)',
            r'^([a-zA-Z0-9-_]+)$'
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        raise ValueError(f"Не удалось извлечь ID таблицы из URL: {url}")

    def connect_to_google_sheets(self, sheet_url_or_id: Optional[str] = None):
        try:
            if sheet_url_or_id:
                if 'docs.google.com' in sheet_url_or_id or '/' in sheet_url_or_id:
                    sheet_id = self.extract_sheet_id_from_url(sheet_url_or_id)
                else:
                    sheet_id = sheet_url_or_id
            else:
                sheet_id = self.config.google_sheet_id

            if not sheet_id:
                raise ValueError("ID Google таблицы не указан")

            cred_path = Path(self.config.credentials_path)
            if not cred_path.is_absolute():
                cred_path = Path(self.config_path).parent / cred_path
            if not cred_path.exists():
                raise FileNotFoundError(f"Файл credentials не найден: {cred_path}")

            self.config.credentials_path = str(cred_path)

            if not self.gc:
                scope = [
                    'https://spreadsheets.google.com/feeds',
                    'https://www.googleapis.com/auth/drive'
                ]
                self._google_creds = Credentials.from_service_account_file(
                    self.config.credentials_path,
                    scopes=scope
                )
                self.gc = gspread.authorize(self._google_creds)
                self._drive_service = build('drive', 'v3', credentials=self._google_creds)

            self.google_sheet = self.gc.open_by_key(sheet_id)
            self.logger.info(f"Успешное подключение к Google Таблице: {sheet_id}")
        except Exception as e:
            self.logger.error(f"Ошибка подключения к Google Таблицам: {e}")
            raise

    def update_config(self, **kwargs):
        for key, value in kwargs.items():
            if hasattr(self.config, key):
                setattr(self.config, key, value)

    def get_excel_sheets(self, excel_path: str) -> List[str]:
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False, keep_vba=True, read_only=False)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            self.logger.error(f"Ошибка чтения Excel файла: {e}")
            return []

    def get_google_sheets(self) -> List[str]:
        try:
            if not self.google_sheet:
                return []
            return [sheet.title for sheet in self.google_sheet.worksheets()]
        except Exception as e:
            self.logger.error(f"Ошибка получения списка Google листов: {e}")
            return []

    def download_google_sheet(self, save_path: str, sheet_names: Optional[List[str]] = None,
                              log_callback: Optional[Callable[[str], None]] = None) -> str:
        try:
            if not self.google_sheet:
                raise ValueError("Не подключено к Google таблице")

            self._log("Скачивание Google таблицы...", log_callback)

            file_id = self.google_sheet.id

            if sheet_names:
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                    request = self._drive_service.files().export_media(
                        fileId=file_id,
                        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )

                    fh = io.BytesIO()
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()
                        if status:
                            self._log(f"Скачивание: {int(status.progress() * 100)}%", log_callback)

                    fh.seek(0)
                    tmp_file.write(fh.read())
                    tmp_file_path = tmp_file.name

                wb = openpyxl.load_workbook(tmp_file_path)
                sheets_to_remove = [sheet for sheet in wb.sheetnames if sheet not in sheet_names]
                for sheet_name in sheets_to_remove:
                    wb.remove(wb[sheet_name])

                wb.save(save_path)
                wb.close()
                os.unlink(tmp_file_path)

                self._log(f"✓ Скачаны листы: {', '.join(sheet_names)}", log_callback)
            else:
                request = self._drive_service.files().export_media(
                    fileId=file_id,
                    mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

                fh = io.FileIO(save_path, 'wb')
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    if status:
                        self._log(f"Скачивание: {int(status.progress() * 100)}%", log_callback)

                fh.close()
                self._log("✓ Скачана полная таблица", log_callback)

            self._log(f"💾 Сохранено: {save_path}", log_callback)
            return save_path

        except Exception as e:
            self._log(f"❌ Ошибка скачивания: {e}", log_callback)
            raise

    def process_excel_file(
            self,
            excel_path: str,
            progress_callback: Optional[Callable[[int, int, str], None]] = None,
            log_callback: Optional[Callable[[str], None]] = None
    ):
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel файл не найден: {excel_path}")

            clear_column_cache()

            self._log("Загрузка Excel файла...", log_callback)
            wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
            wb_values = openpyxl.load_workbook(excel_path, data_only=True)
            self._log(
                "Примечание: openpyxl не вычисляет формулы. Значения берутся из последнего сохранения файла.",
                log_callback
            )

            total_sheets = len(self.config.sheet_mapping)
            processed_sheets = 0

            for excel_sheet_name, google_sheet_name in self.config.sheet_mapping.items():
                try:
                    self._log(f"Начало обработки листа: {excel_sheet_name}", log_callback)

                    if excel_sheet_name not in wb_formulas.sheetnames:
                        self._log(f"⚠️ Лист '{excel_sheet_name}' не найден в Excel файле", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue
                    excel_sheet = wb_formulas[excel_sheet_name]
                    excel_sheet_values = None
                    if excel_sheet_name in wb_values.sheetnames:
                        excel_sheet_values = wb_values[excel_sheet_name]
                    else:
                        self._log(
                            f"⚠️ Лист '{excel_sheet_name}' отсутствует в книге значений (data_only=True). Формулы будут вставлены без вычисленных значений",
                            log_callback
                        )

                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue

                    rows_copied = copy_sheet_data(
                        excel_sheet,
                        google_worksheet,
                        self.config.column_mapping,
                        self.config.start_row,
                        log_callback,
                        excel_sheet_values=excel_sheet_values
                    )

                    self._log(
                        f"✓ Лист '{excel_sheet_name}' обработан. Скопировано строк: {rows_copied}",
                        log_callback
                    )

                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)

                except Exception as e:
                    self._log(f"❌ Ошибка при обработке листа '{excel_sheet_name}': {e}", log_callback)
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)

            wb_formulas.close()
            wb_values.close()
            self._log("✓ Обработка завершена", log_callback)
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise

    def process_multiple_excel_files(
            self,
            file_mappings: List[Dict],
            google_sheet_url: str,
            progress_callback: Optional[Callable[[int, int, str], None]] = None,
            log_callback: Optional[Callable[[str], None]] = None
    ):
        try:
            self._log("Подключение к Google Таблицам...", log_callback)
            self.connect_to_google_sheets(google_sheet_url)

            total_mappings = len(file_mappings)
            processed = 0

            for mapping in file_mappings:
                try:
                    clear_column_cache()

                    excel_path = mapping['excel_path']
                    excel_sheet_name = mapping.get('excel_sheet', 'Sheet1')
                    google_sheet_name = mapping['google_sheet']

                    self._log(
                        f"Обработка: {os.path.basename(excel_path)} → {google_sheet_name}",
                        log_callback
                    )

                    if not os.path.exists(excel_path):
                        self._log(f"⚠️ Файл не найден: {excel_path}", log_callback)
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue

                    wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
                    wb_values = openpyxl.load_workbook(excel_path, data_only=True)
                    self._log(
                        "Примечание: openpyxl не вычисляет формулы. Значения берутся из последнего сохранения файла.",
                        log_callback
                    )

                    if excel_sheet_name not in wb_formulas.sheetnames:
                        if wb_formulas.sheetnames:
                            excel_sheet_name = wb_formulas.sheetnames[0]
                            self._log(f"Используется лист: {excel_sheet_name}", log_callback)
                        else:
                            self._log(f"⚠️ В файле нет листов", log_callback)
                            wb_formulas.close()
                            wb_values.close()
                            processed += 1
                            if progress_callback:
                                progress_callback(processed, total_mappings, os.path.basename(excel_path))
                            continue

                    excel_sheet = wb_formulas[excel_sheet_name]
                    excel_sheet_values = None
                    if excel_sheet_name in wb_values.sheetnames:
                        excel_sheet_values = wb_values[excel_sheet_name]
                    else:
                        self._log(
                            f"⚠️ Лист '{excel_sheet_name}' отсутствует в книге значений (data_only=True). Формулы будут вставлены без вычисленных значений",
                            log_callback
                        )

                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        wb_formulas.close()
                        wb_values.close()
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue

                    self.config.column_mapping = mapping.get('column_mapping', {'source': ['A'], 'target': ['A']})
                    self.config.start_row = mapping.get('start_row', 1)

                    rows_copied = copy_sheet_data(
                        excel_sheet,
                        google_worksheet,
                        self.config.column_mapping,
                        self.config.start_row,
                        log_callback,
                        excel_sheet_values=excel_sheet_values
                    )

                    self._log(f"✓ Скопировано строк: {rows_copied}", log_callback)

                    wb_formulas.close()
                    wb_values.close()
                except Exception as e:
                    self._log(f"❌ Ошибка при обработке {mapping.get('excel_path', 'unknown')}: {e}", log_callback)

                processed += 1
                if progress_callback:
                    progress_callback(processed, total_mappings, os.path.basename(mapping.get('excel_path', 'unknown')))

            self._log("✓ Пакетная обработка завершена", log_callback)
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise

    def _log(self, message: str, log_callback: Optional[Callable[[str], None]] = None):
        self.logger.info(message)
        if log_callback:
            log_callback(message)