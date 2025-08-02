#!/usr/bin/env python3
"""
Отладочный скрипт для проверки формул в Excel файле
"""

import openpyxl
import sys
from pathlib import Path


def debug_excel_formulas(excel_path: str, sheet_name: str = None):
    """Детальный анализ формул в Excel файле"""

    print(f"\n{'=' * 60}")
    print(f"АНАЛИЗ ФАЙЛА: {excel_path}")
    print(f"{'=' * 60}\n")

    # Пробуем разные способы загрузки
    print("1. Загрузка с data_only=False (для чтения формул)...")
    wb_formulas = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)

    print("2. Загрузка с data_only=True (для чтения значений)...")
    wb_values = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    # Выбираем лист
    if sheet_name:
        if sheet_name in wb_formulas.sheetnames:
            sheet_f = wb_formulas[sheet_name]
        else:
            print(f"⚠️ Лист '{sheet_name}' не найден в книге с формулами. Используется активный лист.")
            sheet_f = wb_formulas.active
            sheet_name = sheet_f.title

        if sheet_name in wb_values.sheetnames:
            sheet_v = wb_values[sheet_name]
        else:
            print(f"⚠️ Лист '{sheet_name}' не найден в книге со значениями. Кешированные результаты формул недоступны.")
            sheet_v = None
    else:
        sheet_f = wb_formulas.active
        sheet_name = sheet_f.title
        if sheet_name in wb_values.sheetnames:
            sheet_v = wb_values[sheet_name]
        else:
            print(f"⚠️ Лист '{sheet_name}' не найден в книге со значениями. Кешированные результаты формул недоступны.")
            sheet_v = None

    print(f"\nАнализируем лист: {sheet_name}")
    print(f"Максимальная строка: {sheet_f.max_row}")
    print(f"Максимальная колонка: {sheet_f.max_column}")

    # Счетчики
    total_cells = 0
    cells_with_formulas = 0
    cells_with_values = 0
    formulas_found = []

    print(f"\n{'=' * 60}")
    print("ПОИСК ФОРМУЛ В ПЕРВЫХ 200 СТРОКАХ")
    print(f"{'=' * 60}\n")

    # Проверяем первые 200 строк
    for row in range(1, min(sheet_f.max_row + 1, 200)):
        for col in range(1, sheet_f.max_column + 1):
            cell_f = sheet_f.cell(row=row, column=col)
            cell_v = sheet_v.cell(row=row, column=col) if sheet_v else None

            total_cells += 1

            # Детальная проверка ячейки с формулами
            formula_info = check_cell_for_formula(cell_f, row, col)

            if formula_info['has_formula']:
                cells_with_formulas += 1
                formulas_found.append(formula_info)

                # Выводим первые 10 найденных формул
                if len(formulas_found) <= 10:
                    value_repr = cell_v.value if cell_v else None
                    print(f"ФОРМУЛА НАЙДЕНА в {formula_info['coordinate']}:")
                    print(f"  - Формула: {formula_info['formula']}")
                    print(f"  - Значение (из values): {value_repr}")
                    print(f"  - Тип данных: {formula_info['data_type']}")
                    print(f"  - Способ обнаружения: {formula_info['detection_method']}")
                    if cell_v is None or value_repr is None:
                        print("  - ⚠️ Кешированное значение отсутствует")
                    print()

            # Считаем ячейки с данными
            if cell_v is not None and cell_v.value is not None:
                cells_with_values += 1
            elif cell_v is None and cell_f.value is not None:
                cells_with_values += 1

    # Итоговая статистика
    print(f"\n{'=' * 60}")
    print("СТАТИСТИКА")
    print(f"{'=' * 60}")
    print(f"Всего проверено ячеек: {total_cells}")
    print(f"Ячеек с данными: {cells_with_values}")
    print(f"Ячеек с формулами: {cells_with_formulas}")

    if cells_with_formulas == 0:
        print("\n⚠️ НЕ НАЙДЕНО НИ ОДНОЙ ФОРМУЛЫ!")
        print("\nВозможные причины:")
        print("1. Файл был сохранен с опцией 'Только значения'")
        print("2. Формулы были преобразованы в значения")
        print("3. Используется особый формат файла")
        print("4. Формулы находятся в других местах файла")

        # Проверяем конкретные строки
        print(f"\n{'=' * 60}")
        print("ПРОВЕРКА СТРОК 19 и 31 (которые пропускаются)")
        print(f"{'=' * 60}\n")

        for check_row in [19, 31]:
            print(f"Строка {check_row}:")
            has_data = False
            for col in range(1, min(sheet_f.max_column + 1, 26)):  # A-Z
                cell = sheet_f.cell(row=check_row, column=col)
                if cell.value is not None or cell.data_type == 'f':
                    has_data = True
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"  {col_letter}{check_row}: value={repr(cell.value)}, type={cell.data_type}")
            if not has_data:
                print(f"  Строка полностью пустая")
            print()

    wb_formulas.close()
    wb_values.close()


def check_cell_for_formula(cell, row, col):
    """Детальная проверка ячейки на наличие формулы"""

    coordinate = f"{openpyxl.utils.get_column_letter(col)}{row}"
    info = {
        'coordinate': coordinate,
        'has_formula': False,
        'formula': None,
        'data_type': cell.data_type,
        'detection_method': None
    }

    # Метод 1: Прямая проверка value
    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
        info['has_formula'] = True
        info['formula'] = cell.value
        info['detection_method'] = 'cell.value'
        return info

    # Метод 2: Проверка _value
    if hasattr(cell, '_value') and cell._value is not None:
        value_str = str(cell._value)
        if value_str.startswith('='):
            info['has_formula'] = True
            info['formula'] = value_str
            info['detection_method'] = 'cell._value'
            return info

    # Метод 3: Проверка типа данных
    if cell.data_type == 'f':
        info['has_formula'] = True
        info['formula'] = '=FORMULA_EXISTS_BUT_CANNOT_READ'
        info['detection_method'] = 'data_type=f'

        # Пробуем получить формулу другими способами
        if hasattr(cell, 'formula') and cell.formula:
            info['formula'] = f"={cell.formula}" if not str(cell.formula).startswith('=') else cell.formula
            info['detection_method'] = 'cell.formula'
        elif hasattr(cell, '_formula') and cell._formula:
            info['formula'] = f"={cell._formula}" if not str(cell._formula).startswith('=') else cell._formula
            info['detection_method'] = 'cell._formula'

        return info

    return info


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python debug_formulas.py <путь_к_excel_файлу> [имя_листа]")
        sys.exit(1)

    excel_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    if not Path(excel_path).exists():
        print(f"Файл не найден: {excel_path}")
        sys.exit(1)

    debug_excel_formulas(excel_path, sheet_name)
