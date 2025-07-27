"""
Бизнес-логика для копирования данных из Excel в Google Таблицы
"""

import os
import yaml
import logging
from typing import Dict, List, Tuple, Optional, Callable
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import gspread
from google.oauth2.service_account import Credentials


@dataclass
class Config:
    """Конфигурация приложения"""
    excel_path: str
    google_sheet_id: str
    credentials_path: str
    sheet_mapping: Dict[str, str]
    column_mapping: Dict[str, List[str]]
    start_row: int = 1


class ExcelToGoogleSheets:
    """Класс для копирования данных из Excel в Google Таблицы"""
    
    def __init__(self, config_path: str = "config.yaml"):
        self.config_path = config_path
        self.config = self._load_config(config_path)
        self.logger = self._setup_logger()
        self.gc = None
        self.google_sheet = None
        self._google_creds = None
        
    def _setup_logger(self) -> logging.Logger:
        """Настройка логгера"""
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            
        return logger
    
    def _load_config(self, config_path: str) -> Config:
        """Загрузка конфигурации из YAML файла

        Если файл отсутствует или пустой, возвращается конфигурация
        по умолчанию. Это позволяет корректно инициализировать класс
        даже при первом запуске, когда config.yaml еще не создан.
        """
        if not os.path.exists(config_path):
            # Файл конфигурации отсутствует. Вернём значения по умолчанию.
            return Config(
                excel_path='',
                google_sheet_id='',
                credentials_path='',
                sheet_mapping={},
                column_mapping={'source': ['A'], 'target': ['A']},
                start_row=1,
            )

        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f) or {}

            return Config(
                excel_path=data.get('excel_path', ''),
                google_sheet_id=data.get('google_sheet_id', ''),
                credentials_path=data.get('credentials_path', ''),
                sheet_mapping=data.get('sheet_mapping', {}),
                column_mapping=data.get('column_mapping', {'source': ['A'], 'target': ['A']}),
                start_row=data.get('start_row', 1),
            )
        except Exception as e:
            raise Exception(f"Ошибка загрузки конфигурации: {e}")
    
    def extract_sheet_id_from_url(self, url: str) -> str:
        """Извлечение ID таблицы из URL Google Sheets"""
        import re
        
        # Паттерны для разных форматов URL Google Sheets
        patterns = [
            r'/spreadsheets/d/([a-zA-Z0-9-_]+)',  # Основной паттерн
            r'id=([a-zA-Z0-9-_]+)',                # Альтернативный паттерн
            r'^([a-zA-Z0-9-_]+)$'                  # Просто ID
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        raise ValueError(f"Не удалось извлечь ID таблицы из URL: {url}")
    
    def connect_to_google_sheets(self, sheet_url_or_id: Optional[str] = None):
        """
        Подключение к Google Таблицам
        
        Args:
            sheet_url_or_id: URL или ID Google таблицы (если не указан, берется из конфига)
        """
        try:
            # Определение ID таблицы
            if sheet_url_or_id:
                # Извлекаем ID из URL если передана ссылка
                if 'docs.google.com' in sheet_url_or_id or '/' in sheet_url_or_id:
                    sheet_id = self.extract_sheet_id_from_url(sheet_url_or_id)
                else:
                    sheet_id = sheet_url_or_id
            else:
                sheet_id = self.config.google_sheet_id
                
            if not sheet_id:
                raise ValueError("ID Google таблицы не указан")
            
            # Проверка существования файла credentials
            if not os.path.exists(self.config.credentials_path):
                raise FileNotFoundError(f"Файл credentials не найден: {self.config.credentials_path}")
            
            # Авторизация (только если еще не авторизованы)
            if not self.gc:
                scope = ['https://spreadsheets.google.com/feeds',
                         'https://www.googleapis.com/auth/drive']
                
                self._google_creds = Credentials.from_service_account_file(
                    self.config.credentials_path, 
                    scopes=scope
                )
                
                self.gc = gspread.authorize(self._google_creds)
            
            # Открытие таблицы
            self.google_sheet = self.gc.open_by_key(sheet_id)
            
            self.logger.info(f"Успешное подключение к Google Таблице: {sheet_id}")
            
        except Exception as e:
            self.logger.error(f"Ошибка подключения к Google Таблицам: {e}")
            raise
    
    def update_config(self, **kwargs):
        """Обновление конфигурации"""
        for key, value in kwargs.items():
            if hasattr(self.config, key):
                setattr(self.config, key, value)
    
    def get_excel_sheets(self, excel_path: str) -> List[str]:
        """Получение списка листов из Excel файла"""
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            self.logger.error(f"Ошибка чтения Excel файла: {e}")
            return []
    
    def get_google_sheets(self) -> List[str]:
        """Получение списка листов из Google Таблицы"""
        try:
            if not self.google_sheet:
                return []
            return [sheet.title for sheet in self.google_sheet.worksheets()]
        except Exception as e:
            self.logger.error(f"Ошибка получения списка Google листов: {e}")
            return []
    
    def process_excel_file(self, 
                          excel_path: str, 
                          progress_callback: Optional[Callable[[int, int, str], None]] = None,
                          log_callback: Optional[Callable[[str], None]] = None):
        """
        Обработка Excel файла и копирование данных в Google Таблицы
        
        Args:
            excel_path: Путь к Excel файлу
            progress_callback: Функция обратного вызова для обновления прогресса (current, total, sheet_name)
            log_callback: Функция обратного вызова для логирования
        """
        try:
            # Проверка существования файла
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel файл не найден: {excel_path}")
            
            # Загрузка Excel файла
            self._log("Загрузка Excel файла...", log_callback)
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            
            total_sheets = len(self.config.sheet_mapping)
            processed_sheets = 0
            
            # Обработка каждого листа согласно маппингу
            for excel_sheet_name, google_sheet_name in self.config.sheet_mapping.items():
                try:
                    self._log(f"Начало обработки листа: {excel_sheet_name}", log_callback)
                    
                    # Проверка существования листа в Excel
                    if excel_sheet_name not in wb.sheetnames:
                        self._log(f"⚠️ Лист '{excel_sheet_name}' не найден в Excel файле", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue
                    
                    # Получение листов
                    excel_sheet = wb[excel_sheet_name]
                    
                    # Проверка существования листа в Google Таблицах
                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue
                    
                    # Копирование данных
                    rows_copied = self._copy_sheet_data(
                        excel_sheet, 
                        google_worksheet,
                        log_callback
                    )
                    
                    self._log(f"✓ Лист '{excel_sheet_name}' обработан. Скопировано строк: {rows_copied}", log_callback)
                    
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                    
                except Exception as e:
                    self._log(f"❌ Ошибка при обработке листа '{excel_sheet_name}': {e}", log_callback)
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)
            
            wb.close()
            self._log("✓ Обработка завершена", log_callback)
            
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise
    
    def process_multiple_excel_files(self,
                                   file_mappings: List[Dict],
                                   google_sheet_url: str,
                                   progress_callback: Optional[Callable[[int, int, str], None]] = None,
                                   log_callback: Optional[Callable[[str], None]] = None):
        """
        Обработка нескольких Excel файлов с индивидуальным маппингом
        
        Args:
            file_mappings: Список маппингов вида:
                [{
                    'excel_path': 'file1.xlsx',
                    'excel_sheet': 'Sheet1',
                    'google_sheet': 'Лист1',
                    'column_mapping': {'source': ['A', 'B'], 'target': ['C', 'D']},
                    'start_row': 2
                }, ...]
            google_sheet_url: URL Google таблицы
            progress_callback: Функция обратного вызова для прогресса
            log_callback: Функция обратного вызова для логирования
        """
        try:
            # Подключение к Google Таблицам
            self._log("Подключение к Google Таблицам...", log_callback)
            self.connect_to_google_sheets(google_sheet_url)
            
            total_mappings = len(file_mappings)
            processed = 0
            
            for mapping in file_mappings:
                try:
                    excel_path = mapping['excel_path']
                    excel_sheet_name = mapping.get('excel_sheet', 'Sheet1')
                    google_sheet_name = mapping['google_sheet']
                    
                    self._log(f"Обработка: {os.path.basename(excel_path)} → {google_sheet_name}", log_callback)
                    
                    # Проверка существования файла
                    if not os.path.exists(excel_path):
                        self._log(f"⚠️ Файл не найден: {excel_path}", log_callback)
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue
                    
                    # Загрузка Excel файла
                    wb = openpyxl.load_workbook(excel_path, read_only=True)
                    
                    # Проверка листа в Excel
                    if excel_sheet_name not in wb.sheetnames:
                        # Если указанного листа нет, берем первый
                        if wb.sheetnames:
                            excel_sheet_name = wb.sheetnames[0]
                            self._log(f"Используется лист: {excel_sheet_name}", log_callback)
                        else:
                            self._log(f"⚠️ В файле нет листов", log_callback)
                            wb.close()
                            processed += 1
                            if progress_callback:
                                progress_callback(processed, total_mappings, os.path.basename(excel_path))
                            continue
                    
                    excel_sheet = wb[excel_sheet_name]
                    
                    # Проверка листа в Google
                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        wb.close()
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue
                    
                    # Обновление конфигурации для этого файла
                    self.config.column_mapping = mapping.get('column_mapping', {'source': ['A'], 'target': ['A']})
                    self.config.start_row = mapping.get('start_row', 1)
                    
                    # Копирование данных
                    rows_copied = self._copy_sheet_data(excel_sheet, google_worksheet, log_callback)
                    
                    self._log(f"✓ Скопировано строк: {rows_copied}", log_callback)
                    
                    wb.close()
                    
                except Exception as e:
                    self._log(f"❌ Ошибка при обработке {mapping.get('excel_path', 'unknown')}: {e}", log_callback)
                
                processed += 1
                if progress_callback:
                    progress_callback(processed, total_mappings, os.path.basename(mapping.get('excel_path', 'unknown')))
            
            self._log("✓ Пакетная обработка завершена", log_callback)
            
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise
    
    def _copy_sheet_data(self, excel_sheet, google_worksheet, log_callback=None) -> int:
        """Копирование данных из Excel листа в Google лист"""
        
        source_cols = self.config.column_mapping['source']
        target_cols = self.config.column_mapping['target']
        
        if len(source_cols) != len(target_cols):
            raise ValueError("Количество исходных и целевых колонок должно совпадать")
        
        # Получение данных из Excel
        excel_data = []
        max_row = excel_sheet.max_row
        
        for row_idx in range(self.config.start_row, max_row + 1):
            row_data = []
            has_data = False
            
            for source_col in source_cols:
                cell_value = excel_sheet[f"{source_col}{row_idx}"].value
                if cell_value is not None:
                    has_data = True
                row_data.append(cell_value if cell_value is not None else '')
            
            if has_data:
                excel_data.append(row_data)
        
        if not excel_data:
            self._log("Нет данных для копирования", log_callback)
            return 0
        
        # Подготовка данных для batch update
        updates = []
        
        for row_offset, row_data in enumerate(excel_data):
            google_row = self.config.start_row + row_offset
            
            for col_idx, (value, target_col) in enumerate(zip(row_data, target_cols)):
                cell_address = f"{target_col}{google_row}"
                updates.append({
                    'range': cell_address,
                    'values': [[value]]
                })
        
        # Batch update
        if updates:
            try:
                # Разбиваем на батчи по 1000 ячеек для оптимизации
                batch_size = 1000
                for i in range(0, len(updates), batch_size):
                    batch = updates[i:i + batch_size]
                    google_worksheet.batch_update(batch, value_input_option='USER_ENTERED')
                
                self._log(f"Обновлено ячеек: {len(updates)}", log_callback)
                
            except Exception as e:
                self._log(f"Ошибка при обновлении данных: {e}", log_callback)
                raise
        
        return len(excel_data)
    
    def _log(self, message: str, log_callback: Optional[Callable[[str], None]] = None):
        """Логирование с поддержкой callback"""
        self.logger.info(message)
        if log_callback:
            log_callback(message)


def create_sample_config(path: str = "config.yaml"):
    """Создание примера конфигурационного файла"""
    sample_config = {
        'credentials_path': 'credentials.json',
        'sheet_mapping': {
            'Sheet1': 'Лист1',
            'Sheet2': 'Лист2'
        },
        'column_mapping': {
            'source': ['A', 'C', 'E'],
            'target': ['B', 'D', 'F']
        },
        'start_row': 2
    }
    
    with open(path, 'w', encoding='utf-8') as f:
        yaml.dump(sample_config, f, allow_unicode=True, default_flow_style=False)
    
    print(f"Создан пример конфигурационного файла: {path}")


if __name__ == "__main__":
    # Создание примера конфигурации, если файл не существует
    if not os.path.exists("config.yaml"):
        create_sample_config()
