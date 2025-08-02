from typing import List, Dict, Optional, Callable
import logging

import gspread
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font


def resolve_excel_columns(sheet, columns: List[str]) -> List[str]:
    if not hasattr(resolve_excel_columns, '_header_cache'):
        resolve_excel_columns._header_cache = {}

    sheet_id = id(sheet)
    if sheet_id not in resolve_excel_columns._header_cache:
        header_map = {}
        for cell in sheet[1]:
            if cell.value is not None:
                header_map[str(cell.value).strip().lower()] = cell.column_letter
        resolve_excel_columns._header_cache[sheet_id] = header_map

    header_map = resolve_excel_columns._header_cache[sheet_id]
    result: List[str] = []

    for col in columns:
        col_str = str(col).strip()
        if not col_str:
            continue

        range_found = False
        for delim in ("-", ":"):
            if delim in col_str:
                start, end = col_str.split(delim, 1)
                start_letter = resolve_excel_columns(sheet, [start])[0]
                end_letter = resolve_excel_columns(sheet, [end])[0]
                start_idx = column_index_from_string(start_letter)
                end_idx = column_index_from_string(end_letter)
                step = 1 if end_idx >= start_idx else -1
                for idx in range(start_idx, end_idx + step, step):
                    result.append(get_column_letter(idx))
                range_found = True
                break

        if not range_found:
            if col_str.isdigit():
                result.append(get_column_letter(int(col_str)))
            elif col_str.isalpha():
                result.append(col_str.upper())
            else:
                key = col_str.lower()
                if key in header_map:
                    result.append(header_map[key])
                else:
                    raise ValueError(f"Заголовок '{col}' не найден в Excel листе")
    return result


def resolve_google_columns(worksheet, columns: List[str]) -> List[str]:
    if not hasattr(resolve_google_columns, '_header_cache'):
        resolve_google_columns._header_cache = {}

    sheet_id = id(worksheet)
    if sheet_id not in resolve_google_columns._header_cache:
        headers = worksheet.row_values(1)
        header_map = {}
        for i, val in enumerate(headers):
            if val:
                header_map[str(val).strip().lower()] = get_column_letter(i + 1)
        resolve_google_columns._header_cache[sheet_id] = header_map

    header_map = resolve_google_columns._header_cache[sheet_id]
    result: List[str] = []

    for col in columns:
        col_str = str(col).strip()
        if not col_str:
            continue

        range_found = False
        for delim in ("-", ":"):
            if delim in col_str:
                start, end = col_str.split(delim, 1)
                start_letter = resolve_google_columns(worksheet, [start])[0]
                end_letter = resolve_google_columns(worksheet, [end])[0]
                start_idx = column_index_from_string(start_letter)
                end_idx = column_index_from_string(end_letter)
                step = 1 if end_idx >= start_idx else -1
                for idx in range(start_idx, end_idx + step, step):
                    result.append(get_column_letter(idx))
                range_found = True
                break

        if not range_found:
            if col_str.isdigit():
                result.append(get_column_letter(int(col_str)))
            elif col_str.isalpha():
                result.append(col_str.upper())
            else:
                key = col_str.lower()
                if key in header_map:
                    result.append(header_map[key])
                else:
                    raise ValueError(f"Заголовок '{col}' не найден в Google листе")
    return result


def rgb_to_hex(rgb_color):
    if not rgb_color or rgb_color == "00000000":
        return None

    if len(rgb_color) == 8:
        rgb_color = rgb_color[2:]

    if not rgb_color.startswith('#'):
        rgb_color = '#' + rgb_color

    return rgb_color


def get_cell_formatting(cell):
    formatting = {}

    try:
        if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
            bg_color = rgb_to_hex(str(cell.fill.fgColor.rgb))
            if bg_color and bg_color != '#FFFFFF':
                formatting['backgroundColor'] = {
                    'red': int(bg_color[1:3], 16) / 255,
                    'green': int(bg_color[3:5], 16) / 255,
                    'blue': int(bg_color[5:7], 16) / 255
                }

        if cell.font:
            font_format = {}

            if cell.font.color and hasattr(cell.font.color, 'rgb'):
                text_color = rgb_to_hex(str(cell.font.color.rgb))
                if text_color and text_color != '#000000':
                    font_format['foregroundColor'] = {
                        'red': int(text_color[1:3], 16) / 255,
                        'green': int(text_color[3:5], 16) / 255,
                        'blue': int(text_color[5:7], 16) / 255
                    }

            if cell.font.bold:
                font_format['bold'] = True

            if cell.font.italic:
                font_format['italic'] = True

            if cell.font.size:
                font_format['fontSize'] = int(cell.font.size)

            if font_format:
                formatting['textFormat'] = font_format

        # Выравнивание должно быть на уровне ячейки, а не в textFormat
        if cell.alignment:
            if cell.alignment.horizontal:
                alignment_map = {
                    'left': 'LEFT',
                    'center': 'CENTER',
                    'right': 'RIGHT'
                }
                h_align = alignment_map.get(cell.alignment.horizontal)
                if h_align:
                    formatting['horizontalAlignment'] = h_align

            if cell.alignment.vertical:
                v_alignment_map = {
                    'top': 'TOP',
                    'center': 'MIDDLE',
                    'bottom': 'BOTTOM'
                }
                v_align = v_alignment_map.get(cell.alignment.vertical)
                if v_align:
                    formatting['verticalAlignment'] = v_align

    except Exception as e:
        pass

    return formatting if formatting else None


def convert_excel_formula_to_google(formula: str) -> str:
    if not formula or not formula.startswith('='):
        return formula

    replacements = {
        'ПРОПИСН': 'UPPER',
        'СТРОЧН': 'LOWER',
        'ПРОПНАЧ': 'PROPER',
        'СЦЕПИТЬ': 'CONCATENATE',
        'ДЛСТР': 'LEN',
        'ЛЕВСИМВ': 'LEFT',
        'ПРАВСИМВ': 'RIGHT',
        'ПСТР': 'MID',
        'НАЙТИ': 'FIND',
        'ЗАМЕНИТЬ': 'SUBSTITUTE',
        'ОКРУГЛ': 'ROUND',
        'ЦЕЛОЕ': 'INT',
        'СУММ': 'SUM',
        'СРЗНАЧ': 'AVERAGE',
        'СЧЁТ': 'COUNT',
        'СЧЁТЗ': 'COUNTA',
        'МАКС': 'MAX',
        'МИН': 'MIN',
        'ЕСЛИ': 'IF',
        'И': 'AND',
        'ИЛИ': 'OR',
        'НЕ': 'NOT',
        'ВПЕР': 'VLOOKUP',
        'ГПР': 'HLOOKUP',
        'ИНДЕКС': 'INDEX',
        'ПОИСКПOZ': 'MATCH'
    }

    original_formula = formula
    converted_formula = formula
    for excel_func, google_func in replacements.items():
        converted_formula = converted_formula.replace(excel_func, google_func)

    return converted_formula


def get_cell_formula_simple(cell):
    """Try to extract an Excel formula from an ``openpyxl`` cell.

    Different versions of ``openpyxl`` may store the original formula in
    various attributes.  The previous implementation manually checked a few
    of them which still missed some cases (for example when ``cell.formula``
    is an object rather than a string).  As a result the application treated
    such cells as empty which caused rows containing formulas to be skipped
    and subsequently shifted in Google Sheets.

    This version iterates over all known attributes that may contain a
    formula and normalises the result to a string starting with ``=``.  If a
    formula is detected but its text cannot be retrieved, a placeholder is
    returned so that the row is still considered to contain data.
    """

    for attr in ("value", "_value", "formula", "_formula"):
        val = getattr(cell, attr, None)
        if not val:
            continue
        # ``cell.formula`` may be a specialised object; convert to string
        val_str = str(val)
        if val_str.startswith("="):
            return val_str
        if attr in {"formula", "_formula"}:
            return f"={val_str}"

    if getattr(cell, "data_type", None) == "f":
        # Formula exists but ``openpyxl`` couldn't expose its text
        return "=FORMULA_EXISTS_BUT_CANNOT_READ"

    return None


def copy_sheet_data(
        excel_sheet,
        google_worksheet,
        column_mapping: Dict[str, List[str]],
        start_row: int,
        log_callback: Optional[Callable[[str], None]] = None,
        excel_sheet_values=None
) -> int:
    """Copy data from an Excel sheet to a Google worksheet.

    Parameters
    ----------
    excel_sheet: ``openpyxl`` worksheet loaded with ``data_only=False`` so
        formula objects are preserved.  Passing a sheet from a workbook loaded
        with ``data_only=True`` would strip formulas which in turn breaks
        detection logic.  The function explicitly checks this to avoid
        mistakes.
    excel_sheet_values: optional worksheet from the same workbook loaded with
        ``data_only=True`` used only for retrieving calculated values.
    """

    # Ensure the worksheets come from the expected workbooks
    if getattr(excel_sheet.parent, "data_only", False):
        raise ValueError("excel_sheet must come from a workbook loaded with data_only=False")
    if excel_sheet_values is not None and not getattr(excel_sheet_values.parent, "data_only", False):
        raise ValueError("excel_sheet_values must come from a workbook loaded with data_only=True")

    if log_callback:
        log_callback("Подготовка данных...")
        if excel_sheet_values is None:
            log_callback("⚠️ Закешированные значения формул недоступны. Будут использованы только формулы из Excel.")

    source_cols = resolve_excel_columns(excel_sheet, column_mapping['source'])
    target_cols = resolve_google_columns(google_worksheet, column_mapping['target'])

    if len(source_cols) != len(target_cols):
        raise ValueError("Количество исходных и целевых колонок должно совпадать")

    max_row = excel_sheet.max_row

    if log_callback:
        log_callback("Анализ видимых строк и данных Excel...")
        log_callback(f"📊 Общее количество строк в листе: {max_row}")
        log_callback(f"📊 Начинаем с строки: {start_row}")
        log_callback(f"📊 Строк для проверки: {max_row - start_row + 1}")

        # Быстрая проверка на наличие формул
        formulas_found = 0
        for row in range(start_row, min(start_row + 50, max_row + 1)):
            for col_letter in source_cols[:3]:  # Проверяем первые 3 колонки
                formula_cell = excel_sheet[f"{col_letter}{row}"]
                if hasattr(formula_cell, 'data_type') and formula_cell.data_type == 'f':
                    formulas_found += 1
                    if formulas_found == 1:
                        log_callback(f"🔍 Найдена ячейка с типом 'f' в {col_letter}{row}")
                        log_callback(f"   value: {repr(formula_cell.value)}")
                        if hasattr(formula_cell, '_value'):
                            log_callback(f"   _value: {repr(formula_cell._value)}")
                        break
            if formulas_found > 0:
                break

    if max_row < start_row:
        if log_callback:
            log_callback("Нет данных для копирования")
        return 0

    rows_data = []
    skipped_rows = []  # Для отчета о пропущенных строках
    rows_with_values = 0
    missing_formula_cache = 0

    for row_num in range(start_row, max_row + 1):
        row_dimension = excel_sheet.row_dimensions.get(row_num)
        is_hidden = bool(row_dimension and row_dimension.hidden)

        has_data = False
        row_cells = []

        for col_letter in source_cols:
            # Always take cells from the ``data_only=False`` sheet so that
            # formula objects are available.
            formula_cell = excel_sheet[f"{col_letter}{row_num}"]

            # Получаем формулу
            cell_formula = get_cell_formula_simple(formula_cell)

            # Получаем значение ячейки (может быть из ``data_only=True`` книги)
            if excel_sheet_values is not None:
                value_cell = excel_sheet_values[f"{col_letter}{row_num}"]
                cell_value = value_cell.value
                if cell_formula is not None and cell_value is None:
                    missing_formula_cache += 1
            else:
                cell_value = formula_cell.value

            # Проверяем есть ли РЕАЛЬНЫЕ данные (не пустые ячейки)
            if cell_formula is not None:
                has_data = True
                if log_callback:
                    log_callback(f"📐 Формула в {col_letter}{row_num}: {cell_formula}")
            elif cell_value is not None and str(cell_value).strip() != "":
                has_data = True

            row_cells.append({
                'value': cell_value,
                'formula': cell_formula,
                'formatting': get_cell_formatting(formula_cell)
            })

        if has_data:
            rows_with_values += 1
            rows_data.append(row_cells)
        else:
            reason = "скрытая строка" if is_hidden else "нет данных"
            skipped_rows.append((row_num, reason))

    # Выводим информацию о пропущенных строках
    if skipped_rows and log_callback:
        log_callback(f"⚠️ Пропущено строк: {len(skipped_rows)}")
        for row_num, reason in skipped_rows[:10]:  # Показываем первые 10
            log_callback(f"  - Строка {row_num}: {reason}")
        if len(skipped_rows) > 10:
            log_callback(f"  ... и еще {len(skipped_rows) - 10} строк")

    if log_callback:
        log_callback(f"✅ Найдено {rows_with_values} строк с данными")
        log_callback(
            f"📈 Статистика: всего строк {max_row}, начало с {start_row}, найдено с данными {rows_with_values}")

        rows_with_formulas = 0
        for row_cells in rows_data:
            if any(cell['formula'] is not None for cell in row_cells):
                rows_with_formulas += 1

        if rows_with_formulas > 0:
            log_callback(f"📐 Строк с формулами: {rows_with_formulas}")
        else:
            log_callback("⚠️ Не найдено ни одной строки с формулами!")

        if missing_formula_cache > 0:
            log_callback(
                f"⚠️ Для {missing_formula_cache} формул отсутствуют закешированные значения. "
                "Файл должен быть пересчитан и сохранён в Excel, иначе в Google Sheets будут вставлены только формулы."
            )

    values_to_update = []
    formats_to_apply = []

    for idx, row_cells in enumerate(rows_data):
        current_google_row = start_row + idx
        row_values = []
        row_formats = []

        for j, cell_data in enumerate(row_cells):
            cell_value = cell_data['value']
            cell_formula = cell_data['formula']
            cell_formatting = cell_data['formatting']

            if cell_formula:
                converted_formula = convert_excel_formula_to_google(cell_formula)
                row_values.append(converted_formula)
                if log_callback and current_google_row in [19, 31]:
                    log_callback(
                        f"✓ Записываем формулу в строку {current_google_row}, колонка {target_cols[j]}: {converted_formula}")
            else:
                row_values.append(cell_value if cell_value is not None else '')

            if cell_formatting:
                target_col_index = column_index_from_string(target_cols[j])
                row_formats.append({
                    'row': current_google_row,
                    'col': target_col_index,
                    'format': cell_formatting
                })

        values_to_update.append(row_values)
        formats_to_apply.extend(row_formats)

    if not values_to_update:
        if log_callback:
            log_callback("Нет данных для записи в Google Sheets")
        return 0

    if log_callback:
        log_callback(f"Подготовлено {len(values_to_update)} строк для записи")
        log_callback("Запись данных в Google Sheets...")

    try:
        col_numbers = [column_index_from_string(col) for col in target_cols]
        min_col = min(col_numbers)
        max_col = max(col_numbers)
        start_col_letter = get_column_letter(min_col)
        end_col_letter = get_column_letter(max_col)
        end_row = start_row + len(values_to_update) - 1
        target_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"

        num_cols = max_col - min_col + 1
        index_map = [column_index_from_string(col) - min_col for col in target_cols]

        formatted_values = []
        for row in values_to_update:
            row_values = [''] * num_cols
            for value, idx in zip(row, index_map):
                row_values[idx] = value if value is not None else ''
            formatted_values.append(row_values)

        if log_callback:
            log_callback(f"Обновление диапазона {target_range}...")

        google_worksheet.update(
            target_range,
            formatted_values,
            value_input_option='USER_ENTERED'
        )

        if log_callback:
            log_callback(f"✓ Данные записаны в диапазон {target_range}")

        if formats_to_apply:
            if log_callback:
                log_callback(f"Применение форматирования к {len(formats_to_apply)} ячейкам...")

            try:
                format_requests = []

                for format_data in formats_to_apply:
                    row = format_data['row']
                    col = format_data['col']
                    cell_format = format_data['format']

                    format_request = {
                        'updateCells': {
                            'range': {
                                'sheetId': google_worksheet.id,
                                'startRowIndex': row - 1,
                                'endRowIndex': row,
                                'startColumnIndex': col - 1,
                                'endColumnIndex': col
                            },
                            'rows': [{
                                'values': [{
                                    'userEnteredFormat': cell_format
                                }]
                            }],
                            'fields': 'userEnteredFormat'
                        }
                    }
                    format_requests.append(format_request)

                batch_size = 500  # Увеличено с 100
                import time

                for i in range(0, len(format_requests), batch_size):
                    batch = format_requests[i:i + batch_size]
                    if batch:
                        try:
                            google_worksheet.spreadsheet.batch_update({
                                'requests': batch
                            })

                            if log_callback:
                                log_callback(
                                    f"Применено форматирование к {min(len(batch), len(format_requests) - i)} ячейкам")

                            if i + batch_size < len(format_requests):
                                time.sleep(1)  # 1 секунда между батчами

                        except Exception as batch_error:
                            if "Quota exceeded" in str(batch_error):
                                if log_callback:
                                    log_callback("⚠️ Достигнут лимит API, пропускаем оставшееся форматирование")
                                break
                            else:
                                raise batch_error

            except Exception as e:
                if log_callback:
                    log_callback(f"Предупреждение: не удалось применить все форматирование - {str(e)}")

        if log_callback:
            log_callback(f"✓ Успешно обновлено {len(values_to_update)} строк, {len(target_cols)} колонок")
            if formats_to_apply:
                log_callback(f"✓ Применено форматирование к {len(formats_to_apply)} ячейкам")

        return rows_with_values

    except Exception as e:
        if log_callback:
            log_callback(f"Ошибка при записи в Google Sheets: {e}")
        raise


def clear_column_cache():
    if hasattr(resolve_excel_columns, '_header_cache'):
        resolve_excel_columns._header_cache.clear()
    if hasattr(resolve_google_columns, '_header_cache'):
        resolve_google_columns._header_cache.clear()
