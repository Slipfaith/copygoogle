import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Callable
import time

import gspread
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from google.oauth2.service_account import Credentials

from config import Config, load_config, BASE_DIR


class ExcelToGoogleSheets:
    """Класс для копирования данных из Excel в Google Таблицы."""

    def __init__(self, config_path: str = "config.yaml"):
        path = Path(config_path)
        if not path.is_absolute():
            path = BASE_DIR / path
        self.config_path = str(path)
        self.config = load_config(self.config_path)
        self.logger = self._setup_logger()
        self.gc = None
        self.google_sheet = None
        self._google_creds = None

    def _setup_logger(self) -> logging.Logger:
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        return logger

    def extract_sheet_id_from_url(self, url: str) -> str:
        """Извлечение ID таблицы из URL Google Sheets."""
        import re
        patterns = [
            r'/spreadsheets/d/([a-zA-Z0-9-_]+)',
            r'id=([a-zA-Z0-9-_]+)',
            r'^([a-zA-Z0-9-_]+)$'
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        raise ValueError(f"Не удалось извлечь ID таблицы из URL: {url}")

    def connect_to_google_sheets(self, sheet_url_or_id: Optional[str] = None):
        try:
            if sheet_url_or_id:
                if 'docs.google.com' in sheet_url_or_id or '/' in sheet_url_or_id:
                    sheet_id = self.extract_sheet_id_from_url(sheet_url_or_id)
                else:
                    sheet_id = sheet_url_or_id
            else:
                sheet_id = self.config.google_sheet_id

            if not sheet_id:
                raise ValueError("ID Google таблицы не указан")

            cred_path = Path(self.config.credentials_path)
            if not cred_path.is_absolute():
                cred_path = Path(self.config_path).parent / cred_path
            if not cred_path.exists():
                raise FileNotFoundError(f"Файл credentials не найден: {cred_path}")

            self.config.credentials_path = str(cred_path)

            if not self.gc:
                scope = [
                    'https://spreadsheets.google.com/feeds',
                    'https://www.googleapis.com/auth/drive'
                ]
                self._google_creds = Credentials.from_service_account_file(
                    self.config.credentials_path,
                    scopes=scope
                )
                self.gc = gspread.authorize(self._google_creds)

            self.google_sheet = self.gc.open_by_key(sheet_id)
            self.logger.info(f"Успешное подключение к Google Таблице: {sheet_id}")
        except Exception as e:
            self.logger.error(f"Ошибка подключения к Google Таблицам: {e}")
            raise

    def update_config(self, **kwargs):
        for key, value in kwargs.items():
            if hasattr(self.config, key):
                setattr(self.config, key, value)

    def get_excel_sheets(self, excel_path: str) -> List[str]:
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            self.logger.error(f"Ошибка чтения Excel файла: {e}")
            return []

    def get_google_sheets(self) -> List[str]:
        try:
            if not self.google_sheet:
                return []
            return [sheet.title for sheet in self.google_sheet.worksheets()]
        except Exception as e:
            self.logger.error(f"Ошибка получения списка Google листов: {e}")
            return []

    def _resolve_excel_columns(self, sheet, columns: List[str]) -> List[str]:
        """Преобразование номеров или заголовков Excel в буквы столбцов."""
        result = []

        for col in columns:
            col_str = str(col).strip()
            if not col_str:
                continue

            # Проверяем диапазон вида A-Z
            if '-' in col_str and self._is_column_range(col_str):
                range_cols = self._expand_column_range(col_str)
                result.extend(range_cols)
                continue

            # Проверяем числовой номер колонки
            if col_str.isdigit():
                result.append(get_column_letter(int(col_str)))
                continue

            # Проверяем букву колонки
            if col_str.isalpha() and len(col_str) <= 2:  # A, B, AA, AB и т.д.
                result.append(col_str.upper())
                continue

            # Ищем по заголовку (только если это не похоже на диапазон колонок)
            header_map = {str(cell.value).strip().lower(): cell.column_letter
                          for cell in sheet[1] if cell.value is not None}
            key = col_str.lower()
            if key in header_map:
                result.append(header_map[key])
                continue

            # Если ничего не найдено
            raise ValueError(f"Колонка '{col}' не найдена")

        return result

    def _is_column_range(self, text: str) -> bool:
        """Проверяет, является ли текст диапазоном колонок типа A-Z"""
        if '-' not in text:
            return False
        parts = text.split('-')
        if len(parts) != 2:
            return False
        start, end = parts[0].strip(), parts[1].strip()
        # Проверяем что обе части - это буквы колонок
        return (start.isalpha() and end.isalpha() and
                len(start) <= 2 and len(end) <= 2)

    def _expand_column_range(self, range_text: str) -> List[str]:
        """Расширяет диапазон колонок A-Z в список [A, B, C, ..., Z]"""
        parts = range_text.split('-')
        start_col = parts[0].strip().upper()
        end_col = parts[1].strip().upper()

        # Поддерживаем как одинарные (A-Z), так и двойные буквы (AA-AB)
        start_num = column_index_from_string(start_col)
        end_num = column_index_from_string(end_col)

        if start_num <= end_num:
            return [get_column_letter(i) for i in range(start_num, end_num + 1)]

        raise ValueError(f"Неверный диапазон колонок: {range_text}")

    def _resolve_google_columns(self, worksheet, columns: List[str]) -> List[str]:
        """Преобразование номеров или заголовков Google в буквы столбцов."""
        result = []

        for col in columns:
            col_str = str(col).strip()
            if not col_str:
                continue

            # Проверяем диапазон вида A-Z
            if '-' in col_str and self._is_column_range(col_str):
                range_cols = self._expand_column_range(col_str)
                result.extend(range_cols)
                continue

            # Проверяем числовой номер колонки
            if col_str.isdigit():
                result.append(get_column_letter(int(col_str)))
                continue

            # Проверяем букву колонки
            if col_str.isalpha() and len(col_str) <= 2:
                result.append(col_str.upper())
                continue

            # Ищем по заголовку
            headers = worksheet.row_values(1)
            header_map = {str(val).strip().lower(): get_column_letter(i + 1)
                          for i, val in enumerate(headers) if val}
            key = col_str.lower()
            if key in header_map:
                result.append(header_map[key])
                continue

            # Если ничего не найдено
            raise ValueError(f"Колонка '{col}' не найдена в Google листе")

        return result

    def process_excel_file(
            self,
            excel_path: str,
            progress_callback: Optional[Callable[[int, int, str], None]] = None,
            log_callback: Optional[Callable[[str], None]] = None
    ):
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel файл не найден: {excel_path}")

            self._log("Загрузка Excel файла...", log_callback)
            wb = openpyxl.load_workbook(excel_path, read_only=True)

            total_sheets = len(self.config.sheet_mapping)
            processed_sheets = 0

            for excel_sheet_name, google_sheet_name in self.config.sheet_mapping.items():
                try:
                    self._log(f"Начало обработки листа: {excel_sheet_name}", log_callback)

                    if excel_sheet_name not in wb.sheetnames:
                        self._log(f"⚠️ Лист '{excel_sheet_name}' не найден в Excel файле", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue

                    excel_sheet = wb[excel_sheet_name]

                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        processed_sheets += 1
                        if progress_callback:
                            progress_callback(processed_sheets, total_sheets, excel_sheet_name)
                        continue

                    rows_copied = self._copy_sheet_data_fast(
                        excel_sheet,
                        google_worksheet,
                        log_callback
                    )

                    self._log(
                        f"✓ Лист '{excel_sheet_name}' обработан. Скопировано строк: {rows_copied}",
                        log_callback
                    )

                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)

                except Exception as e:
                    self._log(f"❌ Ошибка при обработке листа '{excel_sheet_name}': {e}", log_callback)
                    processed_sheets += 1
                    if progress_callback:
                        progress_callback(processed_sheets, total_sheets, excel_sheet_name)

            wb.close()
            self._log("✓ Обработка завершена", log_callback)
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise

    def process_multiple_excel_files(
            self,
            file_mappings: List[Dict],
            google_sheet_url: str,
            progress_callback: Optional[Callable[[int, int, str], None]] = None,
            log_callback: Optional[Callable[[str], None]] = None
    ):
        try:
            self._log("Подключение к Google Таблицам...", log_callback)
            self.connect_to_google_sheets(google_sheet_url)

            total_mappings = len(file_mappings)
            processed = 0

            for mapping in file_mappings:
                try:
                    excel_path = mapping['excel_path']
                    excel_sheet_name = mapping.get('excel_sheet', 'Sheet1')
                    google_sheet_name = mapping['google_sheet']

                    self._log(
                        f"Обработка: {os.path.basename(excel_path)} → {google_sheet_name}",
                        log_callback
                    )

                    if not os.path.exists(excel_path):
                        self._log(f"⚠️ Файл не найден: {excel_path}", log_callback)
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue

                    wb = openpyxl.load_workbook(excel_path, read_only=True)

                    if excel_sheet_name not in wb.sheetnames:
                        if wb.sheetnames:
                            excel_sheet_name = wb.sheetnames[0]
                            self._log(f"Используется лист: {excel_sheet_name}", log_callback)
                        else:
                            self._log(f"⚠️ В файле нет листов", log_callback)
                            wb.close()
                            processed += 1
                            if progress_callback:
                                progress_callback(processed, total_mappings, os.path.basename(excel_path))
                            continue

                    excel_sheet = wb[excel_sheet_name]

                    try:
                        google_worksheet = self.google_sheet.worksheet(google_sheet_name)
                    except gspread.exceptions.WorksheetNotFound:
                        self._log(f"⚠️ Лист '{google_sheet_name}' не найден в Google Таблицах", log_callback)
                        wb.close()
                        processed += 1
                        if progress_callback:
                            progress_callback(processed, total_mappings, os.path.basename(excel_path))
                        continue

                    self.config.column_mapping = mapping.get('column_mapping', {'source': ['A'], 'target': ['A']})
                    self.config.start_row = mapping.get('start_row', 1)

                    rows_copied = self._copy_sheet_data_fast(excel_sheet, google_worksheet, log_callback)

                    self._log(f"✓ Скопировано строк: {rows_copied}", log_callback)

                    wb.close()
                except Exception as e:
                    self._log(f"❌ Ошибка при обработке {mapping.get('excel_path', 'unknown')}: {e}", log_callback)

                processed += 1
                if progress_callback:
                    progress_callback(processed, total_mappings, os.path.basename(mapping.get('excel_path', 'unknown')))

            self._log("✓ Пакетная обработка завершена", log_callback)
        except Exception as e:
            self._log(f"❌ Критическая ошибка: {e}", log_callback)
            raise

    def _copy_sheet_data_fast(self, excel_sheet, google_worksheet, log_callback=None) -> int:
        """🚀 БЫСТРОЕ копирование данных - массовая загрузка одним запросом"""
        start_time = time.time()

        source_cols = self._resolve_excel_columns(excel_sheet, self.config.column_mapping['source'])
        target_cols = self._resolve_google_columns(google_worksheet, self.config.column_mapping['target'])

        if len(source_cols) != len(target_cols):
            raise ValueError("Количество исходных и целевых колонок должно совпадать")

        self._log(f"📊 Обработка колонок: {source_cols} → {target_cols}", log_callback)

        # Собираем ВСЕ данные в один большой массив
        excel_data = []
        max_row = excel_sheet.max_row

        self._log(f"📖 Чтение {max_row - self.config.start_row + 1} строк...", log_callback)

        for row_idx in range(self.config.start_row, max_row + 1):
            row_data = []
            has_data = False

            for source_col in source_cols:
                cell_value = excel_sheet[f"{source_col}{row_idx}"].value
                if cell_value is not None:
                    has_data = True
                # Конвертируем None в пустую строку для Google Sheets
                row_data.append(str(cell_value) if cell_value is not None else '')

            if has_data:
                excel_data.append(row_data)

        if not excel_data:
            self._log("⚠️ Нет данных для копирования", log_callback)
            return 0

        self._log(f"📤 Загрузка {len(excel_data)} строк в Google Sheets...", log_callback)

        # МАССОВАЯ ЗАГРУЗКА ОДНИМ ЗАПРОСОМ! 🚀
        try:
            # Определяем диапазон для вставки
            start_col = target_cols[0]
            end_col = target_cols[-1]
            start_row = self.config.start_row
            end_row = start_row + len(excel_data) - 1

            range_name = f"{start_col}{start_row}:{end_col}{end_row}"

            self._log(f"🎯 Диапазон загрузки: {range_name}", log_callback)

            # Одним махом загружаем ВСЕ данные!
            google_worksheet.update(
                range_name,
                excel_data,
                value_input_option='USER_ENTERED'
            )

            elapsed = time.time() - start_time
            self._log(f"⚡ Загрузка завершена за {elapsed:.2f} сек!", log_callback)

        except Exception as e:
            self._log(f"❌ Ошибка при массовой загрузке: {e}", log_callback)

            # Fallback: если массовая загрузка не сработала, используем старый метод
            self._log("🔄 Переключение на порционную загрузку...", log_callback)
            return self._copy_sheet_data_chunked(excel_data, target_cols, google_worksheet, log_callback)

        return len(excel_data)

    def _copy_sheet_data_chunked(self, excel_data, target_cols, google_worksheet, log_callback=None) -> int:
        """📦 Порционная загрузка данных (fallback метод)"""

        chunk_size = 100  # Загружаем по 100 строк за раз
        total_chunks = (len(excel_data) + chunk_size - 1) // chunk_size

        for chunk_idx in range(0, len(excel_data), chunk_size):
            chunk_data = excel_data[chunk_idx:chunk_idx + chunk_size]

            start_col = target_cols[0]
            end_col = target_cols[-1]
            start_row = self.config.start_row + chunk_idx
            end_row = start_row + len(chunk_data) - 1

            range_name = f"{start_col}{start_row}:{end_col}{end_row}"

            try:
                google_worksheet.update(
                    range_name,
                    chunk_data,
                    value_input_option='USER_ENTERED'
                )

                current_chunk = (chunk_idx // chunk_size) + 1
                self._log(f"📦 Chunk {current_chunk}/{total_chunks} загружен", log_callback)

            except Exception as e:
                self._log(f"❌ Ошибка загрузки chunk {current_chunk}: {e}", log_callback)
                raise

        return len(excel_data)

    def _log(self, message: str, log_callback: Optional[Callable[[str], None]] = None):
        self.logger.info(message)
        if log_callback:
            log_callback(message)